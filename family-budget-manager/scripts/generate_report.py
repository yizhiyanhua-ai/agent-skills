#!/usr/bin/env python3
"""
月度报告生成脚本
功能：从 Excel 文件读取数据，生成月度消费报告
"""

import argparse
from datetime import datetime
from typing import Dict, List
from collections import defaultdict

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def load_expenses(excel_path: str, year: int, month: int) -> List[Dict]:
    """从 Excel 文件加载指定月份的消费记录"""
    if not OPENPYXL_AVAILABLE:
        print("⚠️ 需要安装 openpyxl: pip install openpyxl")
        return []
    
    try:
        wb = load_workbook(excel_path)
        ws = wb['消费记录']
        
        expenses = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # 跳过空行
                continue
            
            date_str = row[0] if isinstance(row[0], str) else row[0].strftime('%Y-%m-%d')
            expense_date = datetime.strptime(date_str, '%Y-%m-%d')
            
            if expense_date.year == year and expense_date.month == month:
                expenses.append({
                    'date': date_str,
                    'amount': float(row[1]),
                    'category': row[2],
                    'note': row[3],
                    'payment_method': row[4] if len(row) > 4 else '未指定',
                })
        
        return expenses
    except Exception as e:
        print(f"❌ 读取 Excel 文件失败: {e}")
        return []


def calculate_statistics(expenses: List[Dict]) -> Dict:
    """计算统计数据"""
    if not expenses:
        return {}
    
    total_amount = sum(e['amount'] for e in expenses)
    total_count = len(expenses)
    
    # 按类别统计
    category_stats = defaultdict(lambda: {'amount': 0, 'count': 0})
    for expense in expenses:
        category = expense['category']
        category_stats[category]['amount'] += expense['amount']
        category_stats[category]['count'] += 1
    
    # 计算日均支出
    dates = set(e['date'] for e in expenses)
    days_count = len(dates)
    daily_avg = total_amount / days_count if days_count > 0 else 0
    
    return {
        'total_amount': total_amount,
        'total_count': total_count,
        'daily_avg': daily_avg,
        'category_stats': dict(category_stats),
        'days_count': days_count,
    }


def generate_markdown_report(year: int, month: int, stats: Dict) -> str:
    """生成 Markdown 格式的报告"""
    if not stats:
        return f"# {year}年{month}月消费报告\n\n暂无数据"
    
    report = f"# {year}年{month}月消费报告\n\n"
    
    # 总览
    report += "## 总览\n"
    report += f"- 总支出：¥{stats['total_amount']:.2f}\n"
    report += f"- 记账笔数：{stats['total_count']}笔\n"
    report += f"- 记账天数：{stats['days_count']}天\n"
    report += f"- 日均支出：¥{stats['daily_avg']:.2f}\n\n"
    
    # 类别分布
    report += "## 类别分布\n\n"
    report += "| 类别 | 金额 | 占比 | 笔数 |\n"
    report += "|------|------|------|------|\n"
    
    # 按金额排序
    sorted_categories = sorted(
        stats['category_stats'].items(),
        key=lambda x: x[1]['amount'],
        reverse=True
    )
    
    for category, data in sorted_categories:
        amount = data['amount']
        count = data['count']
        percentage = (amount / stats['total_amount']) * 100
        report += f"| {category} | ¥{amount:.2f} | {percentage:.1f}% | {count} |\n"
    
    report += "\n"
    
    # 消费建议
    report += "## 消费建议\n\n"
    
    # 找出占比最高的类别
    if sorted_categories:
        top_category, top_data = sorted_categories[0]
        top_percentage = (top_data['amount'] / stats['total_amount']) * 100
        
        if top_percentage > 40:
            report += f"- {top_category}支出占比较高（{top_percentage:.1f}%），建议关注是否有优化空间\n"
        
        # 日均支出建议
        if stats['daily_avg'] > 200:
            report += f"- 日均支出较高（¥{stats['daily_avg']:.2f}），建议设置预算控制\n"
        elif stats['daily_avg'] < 50:
            report += f"- 日均支出较低（¥{stats['daily_avg']:.2f}），消费较为节制\n"
    
    return report


def main():
    parser = argparse.ArgumentParser(description='生成月度消费报告')
    parser.add_argument('excel_path', help='Excel 文件路径')
    parser.add_argument('--year', type=int, default=datetime.now().year, help='年份（默认当前年）')
    parser.add_argument('--month', type=int, default=datetime.now().month, help='月份（默认当前月）')
    parser.add_argument('--output', help='输出文件路径（可选，默认打印到终端）')
    
    args = parser.parse_args()
    
    # 加载数据
    expenses = load_expenses(args.excel_path, args.year, args.month)
    
    if not expenses:
        print(f"⚠️ {args.year}年{args.month}月暂无消费记录")
        return
    
    # 计算统计
    stats = calculate_statistics(expenses)
    
    # 生成报告
    report = generate_markdown_report(args.year, args.month, stats)
    
    # 输出
    if args.output:
        with open(args.output, 'w', encoding='utf-8') as f:
            f.write(report)
        print(f"✅ 报告已生成：{args.output}")
    else:
        print(report)


if __name__ == '__main__':
    main()
